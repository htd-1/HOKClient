"""新增 HOK 客户端配置表 client_setting / client_audio。

仿 _rebuild_hok_schema.py 的 openpyxl 直写模式
（luban_helper.py 与本项目 _rebuild_hok_schema 写入的 __tables__ 列结构不兼容，故沿用直写）。

- client_setting：横表 map（单行 id=1），3 个手感 int
- client_audio：横表 map（18 行），key(PascalCase=AudioKey enum 名) / group / path(原 HOKClient 资源名)

音频数据严格来自原 HOKClient AudioSvc 调用点：
  PlayBGMusic / PlayUIAudio / PlayBattleFieldAudio  共 18 条（3 BGM + 7 战斗事件 + 8 UI 音效）。
技能/Buff 音频已在 SkillCfg.AudioStart/Work/Hit、BuffCfg.BuffAudio/HitTickAudio，不在本表。
"""
from pathlib import Path
from openpyxl import Workbook, load_workbook

BASE = Path("E:/Unity.game/HOK/TEngine/Configs/GameConfig/Datas")

# ============ 1. __tables__.xlsx 追加两行 ============
def add_tables():
    p = BASE / "__tables__.xlsx"
    wb = load_workbook(p)
    ws = wb.active
    existing = set()
    for row in ws.iter_rows(min_row=3, values_only=True):
        if row and row[1]:
            existing.add(str(row[1]))
    additions = [
        ("hok.TbClientSetting", "hok.ClientSetting", "id", "map", "c", "HOK客户端手感配置",  "True", "client_setting.xlsx"),
        ("hok.TbClientAudio",   "hok.ClientAudio",   "id", "map", "c", "HOK客户端音频配置",  "True", "client_audio.xlsx"),
    ]
    for full, vt, idx, mode, grp, cmt, rsff, inp in additions:
        if full in existing:
            print(f"  skip (exists): {full}")
            continue
        ws.append(["", full, vt, idx, mode, grp, cmt, rsff, inp, "", ""])
        print(f"  added: {full}")
    wb.save(p)

# ============ 2. client_setting.xlsx（横表 map，单行）============
def write_client_setting():
    wb = Workbook(); ws = wb.active; ws.title = "Sheet1"
    ws.append(["##var",  "id", "opDis",         "skillOPDis",     "skillCancelDis"])
    ws.append(["##type", "int", "int",          "int",            "int"])
    ws.append(["##",     "配置ID", "摇杆触发距离", "技能摇杆距离",   "技能取消距离"])
    ws.append(["##group","c", "c",              "c",              "c"])
    ws.append(["", 1, 135, 125, 500])
    wb.save(BASE / "client_setting.xlsx")
    print("  written: client_setting.xlsx (1 row)")

# ============ 3. client_audio.xlsx（横表 map，18 行）============
# (id, key, group, path, comment)
AUDIO_ROWS = [
    (1,  "MainBgm",           "bgm",        "main",              "主城/登录BGM"),
    (2,  "BattleBgm",         "bgm",        "battle",            "战斗BGM"),
    (3,  "LoadBgm",           "bgm",        "load",              "加载界面BGM"),
    (4,  "Welcombattle",      "sfx_battle", "welcombattle",      "战斗开场"),
    (5,  "Firstblood",        "sfx_battle", "firstblood",        "一血"),
    (6,  "SelfDeath",         "sfx_battle", "selfDeath",         "自己阵亡"),
    (7,  "SelfTowerDestroy",  "sfx_battle", "selfTowerDestroy",  "己方防御塔被毁"),
    (8,  "DestroyEnemyTower", "sfx_battle", "destroyEnemyTower", "摧毁敌方防御塔"),
    (9,  "Victory",           "sfx_battle", "victory",           "胜利"),
    (10, "Defeat",            "sfx_battle", "defeat",            "失败"),
    (11, "LoginBtn",          "sfx_ui",     "loginBtnClick",     "登录按钮"),
    (12, "MatchBtn",          "sfx_ui",     "matchBtnClick",     "匹配/排位/设置按钮"),
    (13, "MatchSureBtn",      "sfx_ui",     "matchSureClick",    "匹配确认"),
    (14, "MatchReminder",     "sfx_ui",     "matchReminder",     "匹配提醒"),
    (15, "SelectHeroBtn",     "sfx_ui",     "selectHeroClick",   "选英雄"),
    (16, "ComClick1",         "sfx_ui",     "com_click1",        "通用点击1"),
    (17, "ComClick2",         "sfx_ui",     "com_click2",        "通用点击2"),
    (18, "ComCdOk",           "sfx_ui",     "com_cd_ok",         "技能CD就绪"),
]

def write_client_audio():
    wb = Workbook(); ws = wb.active; ws.title = "Sheet1"
    ws.append(["##var",  "id", "key",    "group", "path", "comment"])
    ws.append(["##type", "int", "string", "string", "string", "string"])
    ws.append(["##",     "配置ID", "业务键", "分类", "资源名", "注释"])
    ws.append(["##group","c", "c",     "c",     "c",     "c"])
    for row in AUDIO_ROWS:
        ws.append([""] + list(row))
    wb.save(BASE / "client_audio.xlsx")
    print(f"  written: client_audio.xlsx ({len(AUDIO_ROWS)} rows)")

if __name__ == "__main__":
    print("[1/3] __tables__.xlsx")
    add_tables()
    print("[2/3] client_setting.xlsx")
    write_client_setting()
    print("[3/3] client_audio.xlsx")
    write_client_audio()
    print("done.")
