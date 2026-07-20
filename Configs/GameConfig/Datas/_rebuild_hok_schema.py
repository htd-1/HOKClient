"""
重建 HOK Luban schema：清理 TEngine 默认 item/test 示例，
按旧框架硬编码配置定义 8 枚举 + TargetCfg/BulletCfg/SkillCfg/多态BuffCfg bean，
技能与 Buff 数据用 JSON 数据源（多态 $type 分发）。
"""
import json
from pathlib import Path
from openpyxl import Workbook

BASE = Path("E:/Unity.game/HOK/TEngine/Configs/GameConfig/Datas")

# ============ 1. 枚举定义 ============
# (full_name, [(name, value, comment)], flags)
ENUMS = [
    ("hok.ReleaseMode", [("None", 0, "无"), ("Click", 1, "点击施放"), ("Postion", 2, "位置施放"), ("Direction", 3, "方向施放")]),
    ("hok.SelectRule", [("None", 0, ""), ("MinHPValue", 1, "最少总血量"), ("MinHPPercent", 2, "最少百分比血量"),
        ("TargetClosestSingle", 3, "靠近目标单个"), ("PositionClosestSingle", 4, "靠近位置单个"),
        ("TargetClosetMultiple", 5, "靠近目标多个"), ("PositionClosestMultiple", 6, "靠近位置多个"), ("Hero", 7, "所有英雄")]),
    ("hok.TargetTeam", [("Dynamic", 0, "动态"), ("Friend", 1, "友方"), ("Enemy", 2, "敌方")]),
    ("hok.UnitType", [("Hero", 0, "英雄"), ("Soldier", 1, "小兵"), ("Tower", 2, "防御塔")]),
    ("hok.BuffType", [("None", 0, ""), ("HPCure", 1, "治疗"), ("ModifySkill", 2, "技能修改"),
        ("MoveSpeed_Single", 3, "单体加速"), ("ArthurMark", 4, "标记伤害"), ("Silense", 5, "沉默"),
        ("TargetFlashMove", 6, "目标闪现"), ("DirectionFlashMove", 7, "方向闪现"), ("ExecuteDamage", 8, "百分比斩杀"),
        ("Knockup_Group", 9, "群体击飞"), ("Stun_Single_DynamicTime", 10, "动态眩晕"),
        ("HouyiActiveSkillModify", 11, "后羿主动技能修改"), ("Scatter", 12, "散射"),
        ("HouyiPasvAttackSpeed", 13, "后羿被动攻速"), ("HouyiPasvSkillModify", 14, "后羿被动技能修改"),
        ("HouyiPasvMultiArrow", 15, "后羿被动多重射击"), ("HouyiMixedMultiScatter", 16, "后羿混合多重散射"),
        ("MoveSpeed_DynamicGroup", 17, "动态群体移速"), ("MoveSpeed_StaticGroup", 18, "静态群体移速"),
        ("Damage_DynamicGroup", 19, "动态群体伤害"), ("Damage_StaticGroup", 20, "静态群体伤害"), ("MoveAttack", 21, "移动攻击")]),
    ("hok.AttachType", [("None", 0, ""), ("Caster", 1, "施法者"), ("Target", 2, "目标"), ("Indie", 3, "独立"), ("Bullet", 4, "子弹")]),
    ("hok.StaticPosType", [("None", 0, ""), ("SkillCasterPos", 1, "施法者位置"), ("SkillLockTargetPos", 2, "锁定目标位置"), ("BulletHitTargetPos", 3, "子弹命中位置"), ("UIInputPos", 4, "UI输入位置")]),
    ("hok.BulletType", [("UIDirection", 0, "UI方向"), ("UIPosition", 1, "UI位置"), ("SkillTarget", 2, "技能目标"), ("BuffSearch", 3, "Buff搜索")]),
]

# ============ 2. Bean 定义 ============
# (full_name, parent, comment, [(field_name, field_type, comment)])
BEANS = [
    # TargetCfg - 目标配置（嵌套结构，可空使用）
    ("hok.TargetCfg", "", "目标配置", [
        ("targetTeam", "hok.TargetTeam", "目标队伍"),
        ("selectRule", "hok.SelectRule", "选择规则"),
        ("targetTypeArr", "(list#sep=;),hok.UnitType", "目标类型数组"),
        ("selectRange", "float", "查找范围"),
        ("searchDis", "float", "移动攻击搜索距离"),
    ]),
    # BulletCfg - 子弹配置
    ("hok.BulletCfg", "", "子弹配置", [
        ("bulletType", "hok.BulletType", "子弹类型"),
        ("bulletName", "string", "子弹名称"),
        ("resPath", "string", "资源路径"),
        ("bulletSpeed", "float", "速度"),
        ("bulletSize", "float", "大小"),
        ("bulletHeight", "float", "高度"),
        ("bulletOffset", "float", "偏移"),
        ("bulletDelay", "int", "延迟ms"),
        ("canBlock", "bool", "是否可阻挡"),
        ("impacter", "hok.TargetCfg?", "受影响目标"),
        ("bulletDuration", "int", "持续时间ms"),
    ]),
    # SkillCfg - 技能配置
    ("hok.SkillCfg", "", "技能配置", [
        ("id", "int", "配置唯一ID"),
        ("skillID", "int", "技能业务ID"),
        ("iconName", "string", "技能图标"),
        ("aniName", "string", "施法动画"),
        ("releaseMode", "hok.ReleaseMode", "施放方式"),
        ("targetCfg", "hok.TargetCfg?", "目标配置"),
        ("bulletCfg", "hok.BulletCfg?", "弹道配置"),
        ("cdTime", "int", "CD时间ms"),
        ("spellTime", "int", "施法前摇ms"),
        ("isNormalAttack", "bool", "是否普攻"),
        ("skillTime", "int", "技能全长时间ms"),
        ("damage", "int", "基础伤害"),
        ("buffIDArr", "(list#sep=;),int", "附加Buff"),
        ("audioStart", "string", "施法开始音效"),
        ("audioWork", "string", "施法成功音效"),
        ("audioHit", "string", "命中音效"),
    ]),
    # ===== 多态 BuffCfg 基类 =====
    ("hok.BuffCfg", "", "Buff配置基类(多态)", [
        ("id", "int", "配置唯一ID"),
        ("buffID", "int", "Buff业务ID"),
        ("buffName", "string", "Buff名称"),
        ("buffType", "hok.BuffType", "Buff类型"),
        ("attacher", "hok.AttachType", "附着目标"),
        ("impacter", "hok.TargetCfg?", "作用目标"),
        ("buffDelay", "int", "延迟ms"),
        ("buffInterval", "int", "间隔ms"),
        ("buffDuration", "int", "持续时间ms"),
        ("staticPosType", "hok.StaticPosType", "静态位置类型"),
        ("buffAudio", "string", "Buff音效"),
        ("buffEffect", "string", "Buff特效"),
        ("hitTickAudio", "string", "命中音效"),
    ]),
    # ===== BuffCfg 子类（专有字段）=====
    ("hok.CommonBuffCfg", "hok.BuffCfg", "通用Buff", []),
    ("hok.HPCureBuffCfg", "hok.BuffCfg", "治疗Buff", [("cureHPpct", "int", "治疗百分比")]),
    ("hok.MoveSpeedBuffCfg", "hok.BuffCfg", "移速Buff", [("amount", "int", "移速增减百分比")]),
    ("hok.CommonModifySkillBuffCfg", "hok.BuffCfg", "普攻替换Buff", [("originalID", "int", "原技能ID"), ("replaceID", "int", "替换技能ID")]),
    ("hok.ArthurMarkBuffCfg", "hok.BuffCfg", "亚瑟标记Buff", [("damagePct", "int", "伤害百分比")]),
    ("hok.TargetFlashMoveBuffCfg", "hok.BuffCfg", "目标闪现Buff", [("offset", "float", "偏移距离")]),
    ("hok.ExecuteDamageBuffCfg", "hok.BuffCfg", "百分比斩杀Buff", [("damagePct", "int", "伤害百分比")]),
    ("hok.DamageDynamicGroupBuffCfg", "hok.BuffCfg", "动态群体伤害Buff", [("damage", "int", "伤害值")]),
    ("hok.DamageStaticGroupBuffCfg", "hok.BuffCfg", "静态群体伤害Buff", [("damage", "int", "伤害值")]),
    ("hok.StunDynamicTimeBuffCfg", "hok.BuffCfg", "动态眩晕Buff", [("minStunTime", "int", "最小眩晕ms"), ("maxStunTime", "int", "最大眩晕ms")]),
    ("hok.HouyiPasvAttackSpeedBuffCfg", "hok.BuffCfg", "后羿被动攻速Buff", [("overCount", "int", "叠加次数"), ("speedAddtion", "int", "攻速加成"), ("resetTime", "int", "重置时间ms")]),
    ("hok.HouyiMultipleSkillModifyBuffCfg", "hok.BuffCfg", "后羿多重技能修改Buff", [("originalID", "int", "原技能ID"), ("powerID", "int", "强化ID"), ("superPowerID", "int", "超强化ID"), ("triggerOverCount", "int", "触发叠加数"), ("resetTime", "int", "重置时间ms")]),
    ("hok.HouyiScatterSkillModifyBuffCfg", "hok.BuffCfg", "后羿散射技能修改Buff", [("originalID", "int", "原技能ID"), ("powerID", "int", "强化ID"), ("superPowerID", "int", "超强化ID")]),
    ("hok.HouyiScatterArrowBuffCfg", "hok.BuffCfg", "后羿散射箭Buff", [("scatterCount", "int", "散射数量"), ("targetCfg", "hok.TargetCfg?", "散射目标"), ("damagePct", "int", "伤害百分比")]),
    ("hok.HouyiMultipleArrowBuffCfg", "hok.BuffCfg", "后羿多重箭Buff", [("arrowCount", "int", "箭数量"), ("arrowDelay", "int", "箭延迟ms"), ("posOffset", "float", "位置偏移")]),
    ("hok.HouyiMixedMultiScatterBuffCfg", "hok.BuffCfg", "后羿混合多重散射Buff", [("scatterCount", "int", "散射数量"), ("targetCfg", "hok.TargetCfg?", "散射目标"), ("damagePct", "int", "伤害百分比"), ("arrowCount", "int", "箭数量"), ("arrowDelay", "int", "箭延迟ms"), ("posOffset", "float", "位置偏移")]),
]

def write_enums():
    wb = Workbook(); ws = wb.active; ws.title = "Sheet1"
    ws.append(["##var", "full_name", "comment", "flags", "group", "tags", "unique", "*items", "", "", "", ""])
    ws.append(["##var", "", "", "", "", "", "", "name", "alias", "value", "comment", "tags"])
    ws.append(["##", "全名", "注释", "是否位标志", "分组", "标签", "是否唯一", "枚举项", "别名", "值", "注释", "标签"])
    ws.merge_cells(start_row=1, start_column=8, end_row=1, end_column=12)
    for fname, items in ENUMS:
        ws.append(["", fname, "", "False", "c", "", "True", items[0][0], "", items[0][1], items[0][2], ""])
        for nm, val, cmt in items[1:]:
            ws.append(["", "", "", "", "", "", "", nm, "", val, cmt, ""])
    wb.save(BASE / "__enums__.xlsx")

def write_beans():
    wb = Workbook(); ws = wb.active; ws.title = "Sheet1"
    ws.append(["##var", "full_name", "parent", "valueType", "sep", "alias", "comment", "tags", "group", "*fields", "", "", "", "", "", ""])
    ws.append(["##var", "", "", "", "", "", "", "", "", "name", "alias", "type", "group", "comment", "tags", "variants"])
    ws.append(["##", "全名", "父类", "是否值类型", "分隔符", "别名", "注释", "标签", "分组", "字段名", "别名", "类型", "分组", "注释", "标签", "变体"])
    ws.merge_cells(start_row=1, start_column=10, end_row=1, end_column=16)
    for fname, parent, comment, fields in BEANS:
        ws.append(["", fname, parent, "False", "", "", comment, "", "c", "", "", "", "", "", "", ""])
        first = True
        for fn, ft, fc in fields:
            if first:
                row = ws.max_row
                first = False
            else:
                ws.append(["", "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""])
                row = ws.max_row
            ws.cell(row, 10, fn)
            ws.cell(row, 11, "")
            ws.cell(row, 12, ft)
            ws.cell(row, 13, "c")
            ws.cell(row, 14, fc)
            ws.cell(row, 15, "")
            ws.cell(row, 16, "")
    wb.save(BASE / "__beans__.xlsx")

def write_tables():
    wb = Workbook(); ws = wb.active; ws.title = "Sheet1"
    ws.append(["##var", "full_name", "value_type", "index", "mode", "group", "comment", "read_schema_from_file", "input", "output", "tags"])
    ws.append(["##", "全名", "记录类型", "主键", "模式", "分组", "注释", "是否从数据文件读schema", "输入文件", "输出", "标签"])
    tables = [
        ("hok.TbUnit", "hok.Unit", "id", "map", "c", "HOK单位配置", "True", "hok_unit.xlsx"),
        ("hok.TbHero", "hok.Hero", "id", "map", "c", "HOK英雄配置", "True", "hok_hero.xlsx"),
        ("hok.TbMap", "hok.Map", "id", "map", "c", "HOK地图配置", "True", "hok_map.xlsx"),
        ("hok.TbSkill", "hok.SkillCfg", "id", "map", "c", "HOK技能配置", "False", "*@hok_skill.json"),
        ("hok.TbBuff", "hok.BuffCfg", "id", "map", "c", "HOKBuff配置(多态)", "False", "*@hok_buff.json"),
        ("hok.TbTargetRule", "hok.TargetRule", "id", "map", "c", "HOK目标规则配置(占位)", "True", "hok_target_rule.xlsx"),
    ]
    for t in tables:
        ws.append(["", t[0], t[1], t[2], t[3], t[4], t[5], t[6], t[7], "", ""])
    wb.save(BASE / "__tables__.xlsx")

# 删除多余示例表
for f in ["item.xlsx", "hok_bullet.xlsx", "hok_skill.xlsx", "hok_buff.xlsx"]:
    p = BASE / f
    if p.exists(): p.unlink()

write_enums()
write_beans()
write_tables()
print("schema rebuilt: enums/beans/tables written, sample tables removed")
